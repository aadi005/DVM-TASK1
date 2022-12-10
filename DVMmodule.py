def populate(file):
    import openpyxl
    wb_obj = openpyxl.load_workbook(file)
    sh_obj = wb_obj.active
    col = sh_obj.max_column
    row = sh_obj.max_row
    data=[]
    for i in range(1,row+1):
        row_data = []
        for j in range(1,col+1):
            cell_obj = sh_obj.cell(row = i, column = j)
            row_data.append(cell_obj.value)
        data.append(row_data)
    return data
