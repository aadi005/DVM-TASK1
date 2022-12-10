import datetime as dt                           #importing datetime module as dt

from tabulate import tabulate                           #importing tabulate function from tabulate module

import logging                           #importing logging module

from DVMmodule import populate                           #importing populate function from DVMmodule (local module)

import openpyxl                           #importing openpyxl module to work with excel files



logging.basicConfig(filename="DVMLOG.log",format='%(asctime)s %(message)s',filemode='w')                           #Setting up logging file

logger = logging.getLogger()

logger.setLevel(logging.DEBUG)



AllBooks = {}                           #Creating dictionaries to store all the users,books and shelfs

AllUsers = {}

AllShelf = {}



wbBook = openpyxl.Workbook()                           #Setting up working with excel file using openpyxl module

wbUsers = openpyxl.Workbook()

wbShelf = openpyxl.Workbook()



sheetBook = wbBook.active

sheetUser = wbUsers.active

sheetShelf = wbShelf.active





class Book():                           #creating Book class
    def __init__(self,name,author,isbn,owner='library',last=dt.datetime.now(),reserved = "NoOne"):                           #Initializing class
                
	#Initialize class
        
        self.name = name                           #Get name of book
        self.author = author                           #Get author of book
        self.isbn = isbn                           #Get isbn code of book
        self.owner = owner                           #Get the current owner of book
        self.last = last                          #Get the last datetime when the book was updated
        self.reserved = reserved                           #Get who the book is reserved for



    def __str__(self):                           #using string method to print the book
        return f"{self.name} written by {self.author}"


    
    def borrow_book(self,user):                           #creating the borrow function
        # borrow book
        if (self.reserved=="NoOne" or self.reserved.lower()==user.name.lower()) and self.owner.lower()=="library":
            self.owner = user.name
            self.last = dt.datetime.now()
            user.books += [self.name]
            self.reserved="NoOne"
            logger.info("BOOK BORROWED")
        else:
            print("The Book is not available")

            

    def return_book(self):                           #creating return book function
        # return book
        if self.owner.lower()=="library":
            print("Book already in the library")
        self.owner = "Library"
        self.last = dt.datetime.now()
        logger.info("BOOK RETURNED")


        
    def reserve_book(self,user):                           #creating reserve book function
        # reserve book
        self.reserved = user.name





class user():                           #creating the user class
    def __init__(self,name,role):
        self.name = name                           #name of user
        self.role = role                           #role of user
        self.books = []                           #Get all the books the user has read


        
    def __str__(self):
        return f"User Name:{self.name}, User Role:{self.role}"





class Shelf():                           #Creating the shelf class
    def __init__(self,genre,shelf_no,books=[]):                           #initializing the function
        # initialize class
        self.genre = genre                           #genre of the shelf
        self.books = books                           #shelf number
        self.number = shelf_no                           #list of all books in the shelf



    def show_catalog(self):                           
        # show a shelf's catalogue in a nicely formatted way
        catalog = []
        for i in self.books:
            book = [i.name,i.author,i.isbn]
            catalog.append(book)
        print(tabulate(catalog,headers=["BOOK","AUTHOR","ISBN"]))


            
    def add_book(self,book,user):
        
        # add a book to the shelf, make sure only Librarian Users have access to this
        if user.role.lower()=="librarian":
            self.books += [book]
        else:
            print("This function is for librarian only")


		
    def remove_book(self,book,user):
	# remove a book from the shelf, make sure only Librarian Users have access to this
        if user.role.lower()=="librarian":
            if book in self.books:
                self.books.remove(book)
            else:
                print("The book does not exist in the shelf.")
        else:
            print("This function is for librarians only.")



    def get_books_count(self):                           #get total number of books in shelf
        print(f"The total number of books in shelf are {len(self.books)}.")


	
    def populate_book(self,file,user):                           #creating function to populate the shelf
        global AllBooks
        if user.role.lower()=="librarian":
            lst = populate(file) # imported function with necessary arg
            for i in lst:
                book1= Book(i[0],i[1],i[2])
                self.books.append(book1)
                AllBooks[i[2]]= book1
        else:
            print("This function is for librarians only.")





def Get_Users(file):                           #importing user data
    lst = populate(file) # imported function with necessary arg
    users = {}
    for i in lst:
        user1= user(i[0],i[1])
        users[i[0].lower()]= user1
    return users





def Get_Books(file):                           #importing books data
    lst = populate(file) # imported function with necessary arg
    books = {}
    for i in lst:
        book1= Book(i[0],i[1],str(int(i[2])),i[3],i[4],i[5])
        books[str(int(i[2]))]= book1
    return books





def Get_Shelf(file,AllBooks):                           #importing shelf data
    lst = populate(file) # imported function with necessary arg
    shelf = {}
    for i in lst:
        books_ISBN_lst =  str(i[2]).split(";")
        book_lst =[]
        for j in books_ISBN_lst:
            if j in AllBooks.keys():
                book_lst.append(AllBooks[j])
            else:
                logger.error(f"Book with ISBN {j} does not exist.")
        shelf1 = Shelf(i[0],str(int(i[1])),book_lst)
        shelf[str(int(i[1]))] = shelf1
    return shelf





AllBooks = Get_Books("books.xlsx")                           #storing user data

AllUsers = Get_Users("users.xlsx")                           #storing books data

AllShelf = Get_Shelf("shelfs.xlsx",AllBooks)                           #storing shelf data





def menu():                           #Creating menu to get user input
    print("Please select your function:")
    print("1.Borrow a book.")
    print("2.Return a book.")
    print("3.Reserve a book.")
    print("4.Get a list of books read by user.")
    print("5.Show a shelf catalog.")
    print("6.Add book to shelf.")
    print("7.Remove book from shelf")
    print("8.Get book count of a shelf.")
    print("9.Populate a shelf")
    print("10.Quit the system")
    choice = input("Please enter the number corresponding to ur choice:\n")
    while choice not in ['1','2','3','4','5','6','7','8','9','10']:
        choice = input("Please enter the number corresponding to ur choice:\n")
    return int(choice)




#Selecting a user

if len(AllBooks)==0:
    print("CREATE NEW USER!!")
    name = input("Please enter user name:\n")
    role = input("AutodeskLibrarian\n")
    while role.lower() not in ['basic','librarian']:
        role = input("Please choose a role-\n1.Basic\n2.Librarian\n")
    NewUser = user(name,role)
    AllUsers[name.lower()] = NewUser
    Cuser = NewUser



    
else:
    n = 1
    
    for i in AllUsers.keys():
        print(f"{n}.{i}")
        n+=1
    Suser = input("Please Choose a user from the above list or press N to create new User:\n")

    
    while (Suser.lower() not in AllUsers.keys()) and Suser.lower()!="n":
        n = 1
        for i in AllUsers.keys():
            print(f"{n}.{i}")
            n+=1
        Suser = input("Please Choose a user from the above list or press N to create new User:\n")

        
    if Suser.lower()=='n':
        name = input("Please enter user name:\n")
        role = input("Please choose a role-\n1.Basic\n2.Librarian\n")
        
        while role.lower() not in ['basic','librarian']:
            role = input("Please choose a role-\n1.Basic\n2.Librarian\n")
            
        NewUser = user(name,role)
        AllUsers[name.lower()] = NewUser
        Cuser = NewUser

        
    else:
        Cuser =  AllUsers[Suser.lower()]


#Creating a loop so that user may do multiple functions at once

while True:
    choice = menu()                           #calling the menu function and then using if,elif and else accoring to user choice

    
    if choice==1:
        n=1
        lst_book = list(AllBooks.values())
        for i in lst_book:
            print(f"{n}."+str(i))
            n+=1
        choice2 = input("Please enter the number corresponding to your book-\n")
        while choice2 not in [str(i) for i in range(1,len(lst_book)+1)]:
            choice2 = input("Please enter the number corresponding to your book-\n")
        choice2 = int(choice2)
        lst_book[choice2-1].borrow_book(Cuser)

        
    elif choice==2:
        n=1
        lst_book = list(AllBooks.values())
        for i in lst_book:
            print(f"{n}."+str(i))
            n+=1
        choice2 = input("Please enter the number corresponding to your book-\n")
        while choice2 not in [str(i) for i in range(1,len(lst_book)+1)]:
            choice2 = input("Please enter the number corresponding to your book-\n")
        choice2 = int(choice2)
        lst_book[choice2-1].return_book()

        
    elif choice==3:
        n=1
        lst_book = list(AllBooks.values())
        for i in lst_book:
            print(f"{n}."+str(i))
            n+=1
        choice2 = input("Please enter the number corresponding to your book-\n")
        while choice2 not in [str(i) for i in range(1,len(lst_book)+1)]:
            choice2 = input("Please enter the number corresponding to your book-\n")
        choice2 = int(choice2)
        lst_book[choice2-1].reserve_book(Cuser)

        
    elif choice==4:
        n=1
        for i in Cuser.books:
            print(f"{n}."+i)
            n+=1
        print(lst_book[choice2-1].owner)

        
    elif choice==5:
        n=1
        lst_shelf = list(AllShelf.values())
        for i in lst_shelf:
            print(f"{n}."+str(i.genre))
            n+=1
        choice2 = input("Please enter the number corresponding to your shelf-\n")
        while choice2 not in [str(i) for i in range(1,len(lst_shelf)+1)]:
            choice2 = input("Please enter the number corresponding to your shelf-\n")
        choice2 = int(choice2)
        lst_shelf[choice2-1].show_catalog()

        
    elif choice==6:
        name = input("Please Enter Book Name:\n")
        author = input("Please Enter Author Name:\n")
        ISBN = input("Please Enter ISBN:\n")
        while ISBN.isdigit()!=True:
            ISBN = input("Please Enter ISBN:\n")
        
        NewBook = Book(name,author,ISBN)
        AllBooks[ISBN] = NewBook
        n=1
        lst_shelf = list(AllShelf.values())
        for i in lst_shelf:
            print(f"{n}."+str(i.genre))
            n+=1
        choice2 = input("Please enter the number corresponding to your shelf-\n")
        while choice2 not in [str(i) for i in range(1,len(lst_shelf)+1)]:
            choice2 = input("Please enter the number corresponding to your shelf-\n")
        choice2 = int(choice2)
        lst_shelf[choice2-1].add_book(NewBook,Cuser)

        
    elif choice==7:
        ISBN = input("Please Enter ISBN no. of book:\n")
        while ISBN.isdigit()!=True:
            ISBN = input("Please Enter ISBN:\n")
        if ISBN in AllBooks.keys():
            BookRemove= AllBooks[ISBN]
            n=1
            lst_shelf = list(AllShelf.values())
            for i in lst_shelf:
                print(f"{n}."+str(i.genre))
                n+=1
            choice2 = input("Please enter the number corresponding to your shelf-\n")
            while choice2 not in [str(i) for i in range(1,len(lst_shelf)+1)]:
                choice2 = input("Please enter the number corresponding to your shelf-\n")
            choice2 = int(choice2)
            lst_shelf[choice2-1].remove_book(BookRemove,Cuser)
        else:
            print("This Book does not exist.")

            
    elif choice==8:
        n=1
        lst_shelf = list(AllShelf.values())
        for i in lst_shelf:
            print(f"{n}."+str(i.genre))
            n+=1
        choice2 = input("Please enter the number corresponding to your shelf-\n")
        while choice2 not in [str(i) for i in range(1,len(lst_shelf)+1)]:
            choice2 = input("Please enter the number corresponding to your shelf-\n")
        choice2 = int(choice2)
        lst_shelf[choice2-1].get_books_count()

        
    elif choice==9:
        n=1
        lst_shelf = list(AllShelf.values())
        for i in lst_shelf:
            print(f"{n}."+str(i.genre))
            n+=1
        choice2 = input("Please enter the number corresponding to your shelf-\n")
        while choice2 not in [str(i) for i in range(1,len(lst_shelf)+1)]:
            choice2 = input("Please enter the number corresponding to your shelf-\n")
        choice2 = int(choice2)
        lst_shelf[choice2-1].populate_book("DVM.xlsx",Cuser)

        
    else:
        break


#Storing user,book and shelf data back to excel file to save it

n=1

for i in AllBooks.values():
    c1 = sheetBook.cell(row = n, column = 1)
    c2 = sheetBook.cell(row = n, column = 2)
    c3 = sheetBook.cell(row = n, column = 3)
    c4 = sheetBook.cell(row = n, column = 4)
    c5 = sheetBook.cell(row = n, column = 5)
    c6 = sheetBook.cell(row = n, column = 6)
    c1.value = i.name
    c2.value = i.author
    c3.value = i.isbn
    c4.value = i.owner
    c5.value = str(i.last)
    c6.value = i.reserved
    n+=1



n=1
for i in AllUsers.values():
    c1 = sheetUser.cell(row = n, column = 1)
    c2 = sheetUser.cell(row = n, column = 2)
    c1.value = i.name
    c2.value = i.role
    n+=1



n=1
for i in AllShelf.values():
    c1 = sheetShelf.cell(row = n, column = 1)
    c2 = sheetShelf.cell(row = n, column = 2)
    c3 = sheetShelf.cell(row = n, column = 3)
    c1.value = i.genre
    c2.value = i.number
    shelfBooks = ""
    for j in i.books:
        shelfBooks += str(int(j.isbn))
        shelfBooks += ";"
    c3.value = shelfBooks[:-1]
    n+=1



wbBook.save("/Users/aadityagoel/Downloads/DVM TASK 1/books.xlsx")

wbUsers.save("/Users/aadityagoel/Downloads/DVM TASK 1/users.xlsx")

wbShelf.save("/Users/aadityagoel/Downloads/DVM TASK 1/shelfs.xlsx")
        


    

    







'''
BROWNIE CHALLENG-
Test 1- Create new user each of basic and librarian role
Test 2- Use a user with basic role and try and perform each and every function
Test 3- Use a user with librarian role and try and perform each and every function
Test 4- Close the program and run again to check if all the data was saved
'''

'''
External Modules used-
DateTime Module
Tabulate
Logging
Openpyxl
'''




























	
